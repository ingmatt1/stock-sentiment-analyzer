from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import pandas as pd
import xlsxwriter
from nltk.sentiment.vader import SentimentIntensityAnalyzer
import numpy as np

from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference,
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.layout import Layout, ManualLayout
from StyleFrame import StyleFrame, Styler, utils

def sentiment(sentiment_df):
    score_list = []
    
    for index, row in sentiment_df.iterrows():
        score=sentiment_df.loc[index, 'compound']
        
        if score >= 0.05 : 
            score = "positive"
        elif score <= - 0.05 : 
            score = "negative"
        else: 
            score ="neutral"
        
        score_list.append(score)
        
    sentiment_df['score']=score_list
        
    return sentiment_df

def scored_news(tickers):
    finviz = 'https://finviz.com/quote.ashx?t='
    
    news_tables = {}
    
    for ticker in tickers:
        url = finviz + ticker
        req = Request(url=url,headers={'user-agent': 'my-app/0.0.1'}) 
        response = urlopen(req)    
        html = BeautifulSoup(response, features="lxml")
        news_table = html.find(id='news-table')
        news_tables[ticker] = news_table
        
    parsed = []
    
    for name, table in news_tables.items():
        for i in table.findAll('tr'):
            a_text = i.a.get_text() 
            td_text = i.td.text.split()
    
            if len(td_text) == 1:
                time = td_text[0]
            else:
                date = td_text[0]
                time = td_text[1]
    
            parsed.append([name, date, time, a_text])
            
    vader = SentimentIntensityAnalyzer()
    cols = ['ticker', 'date', 'time', 'headline']

    scored_news = pd.DataFrame(parsed, columns=cols)
    scores = scored_news['headline'].apply(vader.polarity_scores).tolist()
    scores_df = pd.DataFrame(scores)

    scored_news = scored_news.join(scores_df, rsuffix='right')
    
    scored_news['date'] = pd.to_datetime(scored_news.date).dt.date
    
    return sentiment(scored_news)

def just_sentiment(tickers): 
    sentiment_df = scored_news(tickers)
    del sentiment_df['pos']
    del sentiment_df['neg']
    del sentiment_df['compound']
    del sentiment_df['neu']
    
    return sentiment_df

def date_range(df):
    date_ranges = {}
    n_tickers=int(len(df.index)/100)
    for i in range(n_tickers):
        if i==0: 
            recent_date = df.loc[i, 'date']
            furthest_date = df.loc[99, 'date']
            ticker=df.loc[i, 'ticker']
        else: 
            recent_date = df.loc[i*100, 'date']
            furthest_date = df.loc[i*100+99, 'date']
            ticker=df.loc[i*100, 'ticker']
        

        date_str = furthest_date.strftime("%b %d, %Y") + ' to ' + recent_date.strftime("%b %d, %Y")
        date_ranges[ticker]=date_str
        
    return date_ranges 

def tally_scores(df): 
    n_tickers=int(len(df.index)/100)
    tally ={}
    tickers=[]
    for i in range(n_tickers):
        if i==0: 
            ticker=df.loc[i, 'ticker']
        else: 
            ticker=df.loc[i*100, 'ticker']
        tickers.append(ticker)
    
    tally['tickers']=tickers
    tally['positive']=np.zeros([n_tickers])
    tally['neutral']=np.zeros([n_tickers])
    tally['negative']=np.zeros([n_tickers])
    tally_df = pd.DataFrame(tally)
    
    ticker_count = 0 
    for i in range(n_tickers*100):
        previous = tally_df.loc[ticker_count, df.loc[i, 'score']]
        tally_df.loc[ticker_count, df.loc[i, 'score']] = previous + 1
        
        if (i+1) % 100 == 0:
            ticker_count = ticker_count+1
            
    return tally_df

def daily_sentiment(tickers):
    df=scored_news(tickers)
    mean_scores = df.groupby(['ticker','date']).mean()
    #optional
    mean_scores = mean_scores.unstack()
    return mean_scores

def plot_sentiment(tickers):
    df=scored_news(tickers)
    
    #getrange of headlines dates and daily scores
    dates = date_range(df)
    
    daily = daily_sentiment(tickers)
    df=tally_scores(df).T
    blank = []

    
    columns=list(df)
    
    path = "C:\\Users\\matth\\Desktop\\stock-sentiment-analyzer\\pie.xlsx"
    sf=StyleFrame(daily.T)
    col_list=list(daily.columns)
    writer = pd.ExcelWriter(path, engine='xlsxwriter')
    daily.T.to_excel(excel_writer=writer, sheet_name = 'Daily Sentiment Summary') 
    worksheet = writer.sheets['Daily Sentiment Summary']                      
    worksheet.set_column('B:B', 15)
    writer.save()
    
    wb = load_workbook("pie.xlsx")
    for i in range(0, len(df.columns)):
        
        ws = wb.create_sheet(tickers[i])
        score=1

        for index, row in df.iterrows():
            vals = []
            if index=='tickers':
                continue
            vals.append(df.loc[index, i])
            vals.insert(0,df.index[score])
            score=score+1
            ws.append(vals)
        
        pie = PieChart()
        labels = Reference(ws, min_col=1, max_col= 1, min_row=1, max_row=3)
        
        data = Reference(ws, min_col=2, max_col = 2, min_row=1, max_row=3)
        pie.add_data(data)
        
        pie.set_categories(labels)
        pie.title = tickers[i] + ' Sentiment'
        pie.layout=Layout(
        manualLayout=ManualLayout(
        h=0.7, w=0.7,
        )
        )
        ws.add_chart(pie, "E3")
        
    
    wb.save("pie.xlsx")
    




#List the tickers of the companies you'd like to analyze
#Amazon, Microsoft, and Google are used as examples
tickers=['AMZN', 'MSFT', 'GOOG']      
plot_sentiment(tickers)









